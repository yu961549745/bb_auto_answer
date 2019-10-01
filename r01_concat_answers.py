""" 拼接原始答案到一个 sheet 
"""
# %%
import re
import pandas as pd
from openpyxl import load_workbook

in_file = '原始答案.xlsx'
# %%
wb = load_workbook(filename=in_file)
sheets = wb.sheetnames
wb.close()


# %%
cols = ['类型', '题干', '备选项A', '备选项B', '备选项C', '备选项D', '试题答案']
ans = dict()
for st in sheets:
    tmp = pd.read_excel(in_file, sheet_name=st).assign(类型=st)
    tmp = tmp[cols]
    ans[st] = tmp

ans = pd.concat(list(ans.values()), axis=0)

# %%


def normalize(s):
    if pd.isna(s):
        return ''
    if isinstance(s, str):
        return re.sub(r'[—_\s]+', '', s)


ans = ans.applymap(normalize)


# %%
ans.to_excel('整理后答案.xlsx', index=False)


# %%
